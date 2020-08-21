import codecs
import json
import os

import openpyxl


class info():
    def __init__(self, file):
        xlsx = openpyxl.load_workbook(file)
        sheet = xlsx.active
        self.data = {}

        for rows in sheet.iter_rows(2, 0, 0, sheet.max_column):
            group = []
            productRevenue = []
            product = []
            concept = []
            on = ''

            if rows[4].value != None:
                group = rows[4].value.split(',')

            if rows[5].value != None:
                productRevenue = rows[5].value.split(',')

            if rows[6].value != None:
                product = rows[6].value.split(',')

            if rows[8].value != None:
                concept = rows[8].value.split(',')

            if rows[11].value == '1':
                on = 'tws'
            else:
                on = 'otc'

            self.data[rows[0].value] = {
                'code': rows[0].value,
                'name': rows[1].value,
                'value': int(rows[2].value) * 1000,
                'industryName': rows[3].value,
                'group': group,
                'productRevenue': productRevenue,
                'product': product,
                'concept': concept,
                'industryPosition': rows[9].value,
                'on': on,
                'close': bool(rows[12].value),
            }

    def output(self, output):
        f = codecs.open(os.path.join(output, 'stock.json'), 'w+', 'utf-8')
        f.write(json.dumps(self.data, ensure_ascii=False))
        f.close()
