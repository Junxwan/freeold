import codecs
import json
import logging
import os

import openpyxl

column = ['open', 'close', 'max', 'min', 'increase', 'amplitude', 'volume']

path = {
    'open': 'price',
    'close': 'price',
    'max': 'price',
    'min': 'price',
    'increase': 'price',
    'amplitude': 'price',
    'volume': 'volume',
}


def run(input, output):
    logging.info('read: ' + input + ' ...')
    xlsx = openpyxl.load_workbook(input)
    sheet = xlsx.active

    row = sheet.cell(1, 3).value
    m = row[:6]
    date = row[:4] + '-' + row[4:6] + '-' + row[6:8]

    data = {}
    for c in column:
        data[c] = []

    for rows in sheet.iter_rows(2, 0, 0, sheet.max_column):
        for index, row in enumerate(rows[2:]):
            data[column[index]].append({
                'code': rows[0].value,
                'name': rows[1].value,
                'value': row.value,
            })

    for name, dName in path.items():
        dir = os.path.join(output, dName, name, m)
        os.makedirs(dir, exist_ok=True)

        filePath = os.path.join(dir, date) + ".json"

        if os.path.exists(filePath):
            continue

        f = codecs.open(filePath, 'w+', 'utf-8')
        f.write(json.dumps({
            'date': date,
            'value': data[name],
        }, ensure_ascii=False))
        f.close()

        logging.info('save ' + name + ' ' + filePath)

    logging.info('total: ' + str(path.__len__()))
