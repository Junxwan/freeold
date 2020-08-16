import codecs
import json
import logging
import os

import openpyxl


def run(input, output):
    paths = []
    if os.path.isfile(input):
        paths.append(input)

    if os.path.isdir(input):
        for f in os.listdir(input):
            fullPath = os.path.join(input, f)

            if os.path.splitext(fullPath)[-1] == '.xlsx':
                paths.append(fullPath)

    for path in paths:
        toJson(path, output)


def toJson(input, output):
    logging.info('read: ' + input + ' ...')
    xlsx = openpyxl.load_workbook(input)
    sheet = xlsx.active
    dates = []
    price = {}

    for rows in sheet.iter_rows(0, 1, 0, sheet.max_column):
        for row in rows[2:]:
            date = row.value[:4] + '-' + row.value[4:6] + '-' + row.value[6:8]
            dates.append(date)

            if price.get(row.value[:6]) == None:
                price[row.value[:6]] = {}

            if price[row.value[:6]].get(date) == None:
                price[row.value[:6]][date] = []

    for rows in sheet.iter_rows(2, 0, 0, sheet.max_column):
        for index, row in enumerate(rows[2:]):
            date = dates[index]
            price[date[:4] + date[5:7]][date].append({
                'code': rows[0].value,
                'name': rows[1].value,
                'value': row.value,
            })

    for name, month in price.items():
        dir = os.path.join(output, name)

        if os.path.exists(dir) == False:
            os.mkdir(dir)

        logging.info('======================== start ' + name + ' ========================')
        logging.info('source xlsx: ' + input)
        logging.info('output path: ' + dir)

        for date, value in month.items():
            path = os.path.join(dir, date) + ".json"

            if os.path.exists(path):
                continue

            f = codecs.open(path, 'w+', 'utf-8')
            f.write(json.dumps({
                'date': date,
                'value': value,
            }, ensure_ascii=False))
            f.close()

            logging.info('date: ' + date)

        logging.info('======================== end ' + name + ' ========================')
