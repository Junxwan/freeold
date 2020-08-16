import codecs
import json
import os

import openpyxl


def run(input, output):
    xlsx = openpyxl.load_workbook(input)
    sheet = xlsx.active
    stock = {}

    for rows in sheet.iter_rows(2, 0, 0, sheet.max_column):
        stock[rows[0].value] = {
            'code': rows[0].value,
            'name': rows[1].value,
            'value': rows[2].value,
            'industryName': rows[3].value,
            'industryCode': rows[4].value,
            'industryIndexName': rows[5].value,
            'industrySubName': rows[6].value,
            'on': rows[7].value,
            'twsDate': rows[8].value,
            'otcDate': rows[9].value,
            'openDate': rows[10].value,
        }

    f = codecs.open(os.path.join(output, 'stock.json'), 'w+', 'utf-8')
    f.write(json.dumps(stock, ensure_ascii=False))
    f.close()
