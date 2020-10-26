# -*- coding: utf-8 -*-

import codecs
import glob
import json
import logging
import os
import time
import openpyxl
import pandas as pd
import numpy as np
from stock import data as dt, name


class Tick():
    def __init__(self, dir, trend_dir):
        self.dir = dir
        self.trend = dt.Trend(trend_dir)

        date_times = pd.date_range(start='2020-01-01 09:00:00', end=f'2020-01-01 13:25:00', freq='min')
        self.times = [t.strftime('%H:%M:%S') for i, t in enumerate(date_times)]

    def output(self, dir):
        for file in sorted(glob.glob(os.path.join(self.dir, "*.xlsx")), reverse=True):
            names = os.path.basename(file).split('.')
            if names[1] != 'xlsx':
                continue

            code = names[0][11:]
            date = names[0][:10]

            path = os.path.join(dir, date)
            if os.path.exists(path) == False:
                os.mkdir(path)

            file_path = os.path.join(path, code) + '.csv'
            if os.path.exists(file_path):
                self.apply_trend(code, date)
                continue

            data = pd.read_excel(file)
            data.columns = ['time', 'buy', 'sell', 'price', 'volume', 'total_volume']
            data = data.reindex(index=data.index[::-1])
            data.insert(len(data.columns), 'amount', (data['price'] * data['volume']).to_numpy().tolist())
            data.insert(len(data.columns), 'avg', np.full([1, len(data)], np.nan).tolist()[0])
            times = data['time']

            # 檢查tick資料跟日內趨勢資料對比
            trend = self.trend.code(code, date)

            if trend is None:
                logging.error(f'tick file data error: {file} not trend')
                return

            column_ary = np.random.randint(len(trend.dropna(axis='columns', how='all').columns), size=10)

            if len(column_ary) == 0:
                logging.error(f'tick file data error: {file} not column trend')
                return

            for i in column_ary:
                d = trend[str(i)]

                if d.empty:
                    continue

                start = d.time[11:]
                end = f'{start[:6]}59'
                v = data[times.between(start, end)]

                if len(v) == 0:
                    continue

                if (v[dt.VOLUME].sum() != int(d[dt.VOLUME])) | (float(d[name.CLOSE]) != v.iloc[-1]['price']):
                    logging.error(f'tick file data error: {file} time: {start}')
                    return

            total_amount = 0
            for index, t in enumerate(self.times[1:]):
                q = data[(self.times[index] <= times) & (times < t)]
                amount = q['amount'].sum()

                if amount == 0:
                    continue

                total_amount += amount

                avg = total_amount / q['total_volume'].iloc[-1]

                if avg < 50:
                    l = 2
                elif avg < 500:
                    l = 1
                else:
                    l = 0

                data['avg'].loc[q.index[0]:q.index[-1]] = round(avg, l)

            data.to_csv(file_path, index=False)
            logging.info(f'tick date:{date} {code}.csv')

            self.apply_trend(code, date)

    def apply_trend(self, code, date):
        code = int(code)
        tick = self.trend.tick(code, date)
        all_trend = self.trend.read(date)
        trend = all_trend.loc[code]

        if trend.loc[name.OPEN].dropna().empty == False & trend.loc[name.AVG].dropna().empty == False:
            return True

        times = [t[11:] for t in trend.loc[name.TIME].dropna()]
        q = tick[name.TIME]
        avg = []
        open = []

        for i, time in enumerate(times):
            d = tick[q.between(times[i - 1], time)]

            if d.empty == False:
                d = d.dropna()
                open.append(d.iloc[0][name.PRICE])
                avg.append(d.iloc[-1][name.AVG])

        [avg.append(np.nan) for _ in range(len(trend.loc[name.TIME]) - len(avg))]
        [open.append(np.nan) for _ in range(len(trend.loc[name.TIME]) - len(open))]

        tmp = all_trend.copy()
        tmp.loc[(int(code), name.AVG), :] = avg
        tmp.loc[(int(code), name.OPEN), :] = open

        tmp.to_csv(self.trend.file_name(date))
        self.trend.data[date] = tmp
        logging.info(f'update tick date:{date} {code}.csv')


# 個股基本資料
class stock():
    def __init__(self, file):
        xlsx = openpyxl.load_workbook(file)
        sheet = xlsx.active
        self.__data = {}

        for rows in sheet.iter_rows(2, 0, 0, sheet.max_column):
            self.__data[rows[0].value] = {
                'code': rows[0].value,
                'name': rows[1].value,
                'value': rows[2].value,
                'industryName': rows[3].value,
                'industryCode': rows[4].value,
                'industryIndexName': rows[5].value,
                'industrySubName': rows[6].value,
                'on': rows[7].value,
                'twsDate': rows[8].value,
                'otcDate': rows[9].value,
                'openDate': rows[10].value,
            }

    def output(self, path):
        f = codecs.open(os.path.join(path, 'cmoney-stock.json'), 'w+', 'utf-8')
        f.write(json.dumps(self.__data, ensure_ascii=False))
        f.close()


# 每日個股行情
class day():
    def __init__(self, file):
        logging.info('read: ' + file + ' ...')

        self.columns = dt.COLUMNS.copy()
        self.columns.insert(0, dt.DATE)
        self.data = pd.DataFrame()
        self.date = os.path.basename(file).split('.')[0]

        for i, rows in pd.read_excel(file).iterrows():
            for ii, value in enumerate(rows[2:]):
                if pd.isna(value):
                    continue

                code = rows[0]

                if code not in self.data:
                    self.data[code] = {c: 0 for c in self.columns}
                    self.data[code][dt.DATE] = self.date

                self.data[code][self.columns[ii + 1]] = value

    def output(self, path):
        values = []
        ck = list(self.data.keys())
        m = f'{self.date[:4]}{self.date[5:7]}'
        filePath = os.path.join(path, m) + ".csv"

        for c in ck:
            if c not in self.data:
                [values.append(np.nan) for _ in self.columns]
            else:
                close = self.data[c][name.CLOSE]
                open = self.data[c][name.OPEN]
                self.data[c][name.D_INCREASE] = round(((close - open) / open) * 100, 2)

                [values.append(v) for n, v in self.data[c].items()]

        index = pd.MultiIndex.from_product([ck, self.columns], names=['code', 'name'])
        data = pd.DataFrame({"0": values}, index=index)

        if os.path.exists(filePath):
            d = pd.read_csv(filePath, index_col=[0, 1], header=[0])
            d.columns = np.arange(data.columns.size, d.columns.size + data.columns.size)
            data = pd.merge(data, d, on=['code', 'name'], how='left')

        data.to_csv(filePath)

        logging.info(f'save file: {filePath}')


# 每年個股行情
class year():
    dataColumns = dt.COLUMNS
    columns = dataColumns.copy()

    def __init__(self, path):
        self.data = {}
        self.path = path
        self.columns.insert(0, dt.DATE)
        years = [os.path.basename(f).split('.')[0] for f in glob.glob(os.path.join(path, dt.OPEN, '*.xlsx'))]
        last = years[-1]

        for year in years:
            for column in self.dataColumns:
                file = os.path.join(path, column, f'{year}.xlsx')

                if os.path.exists(file) == False:
                    continue

                logging.info(f'read: {file}...')

                for i, rows in pd.read_excel(file).iterrows():
                    for ii, value in enumerate(rows[2:]):
                        if pd.isna(value):
                            continue

                        code = rows[0]
                        d = rows.index[ii + 2]
                        date = d[:4] + '-' + d[4:6] + '-' + d[6:8]

                        if year == last:
                            m = d[:6]
                        else:
                            m = year

                        if m not in self.data:
                            self.data[m] = {}

                        if date not in self.data[m]:
                            self.data[m][date] = {}

                        if code not in self.data[m][date]:
                            self.data[m][date][code] = {c: 0 for c in self.columns}
                            self.data[m][date][code][dt.DATE] = date

                        self.data[m][date][code][column] = value

    def output(self, path):
        logging.info('save start')

        for f_name, dates in self.data.items():
            i = 0
            data = {}
            filePath = os.path.join(path, f_name) + ".csv"
            ck = list(dates[list(dates.keys())[0]].keys())
            index = pd.MultiIndex.from_product([ck, self.columns], names=['code', 'name'])

            for date, codes in dates.items():
                values = []

                for c in ck:
                    if c not in codes:
                        [values.append(np.nan) for _ in self.columns]
                    else:
                        close = codes[c][name.CLOSE]
                        open = codes[c][name.OPEN]
                        codes[c][name.D_INCREASE] = round(((close - open) / open) * 100, 2)

                        [values.append(v) for n, v in codes[c].items()]

                data[i] = values
                i += 1

            pd.DataFrame(data, index=index).to_csv(filePath)

            logging.info(f'save {f_name}')

        logging.info('save end')
