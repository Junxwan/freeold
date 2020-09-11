# -*- coding: utf-8 -*-

import logging
import math
import os
import time
import openpyxl
import requests
from bs4 import BeautifulSoup


# 個股基本資料
def pullInfo(input, output):
    wb = openpyxl.Workbook()
    newSheet = wb.active
    xlsx = openpyxl.load_workbook(input)
    sheet = xlsx.active

    for i, code in enumerate(sheet.iter_rows()):
        industryName = ''
        product = []

        code = str(code[0].value)
        resp = requests.get('https://tw.stock.yahoo.com/d/s/company_' + code + '.html')

        if resp.status_code != 200:
            logging.info(code + ' empty')
        else:
            logging.info(code + ' get')

            soup = BeautifulSoup(resp.text, 'html.parser')
            tds = soup.select('table')[2].findAll('td')

            industryName = tds[3].text.strip()
            revenues = tds[31].text.lstrip().split('、')
            revenues[-1] = revenues[-1].partition('%')[0] + '%'

            for i, name in enumerate(revenues):
                for ii, s in enumerate(name):
                    if s.isdigit():
                        r = math.floor(float(name[ii:-1]))
                        if (r >= 1) & (name[:ii] != '其他'):
                            product.append(name[:ii] + '-' + str(r))
                        break

        newSheet.append([code, industryName, ','.join(product)])

        time.sleep(0.5)

        if i % 100 == 0:
            wb.save(os.path.join(output, "yahoo-info.xlsx"))

    wb.save(os.path.join(output, "yahoo-info.xlsx"))
