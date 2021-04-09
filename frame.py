# -*- coding: utf-8 -*-

import glob
import os
import time
import tkinter as tk
import logging
import pandas as pd
import openpyxl
import pyautogui
from stock import name, data
from datetime import datetime
from tkinter import messagebox
from PIL import Image, ImageTk
from ui import cmoney, xq, stock, log, other, ui, watch, execl, twse
from auto import xq as xqa


class main():
    def __init__(self, root, config=None, path=None):
        self.root = root
        self.config = config
        self.size = pyautogui.size()
        self.width = int(self.size.width / 2)
        self.height = int(self.size.height / 2)
        self.w = self.width / 100
        self.h = self.height / 100
        self.isTop = False
        self.currentPath = path

        root.geometry(f'{self.width}x{self.height}')
        self.mainLayout()
        self.frameLayout()

    # 主layout
    def mainLayout(self):
        self.topHeight = int(self.height * 0.6)

        self.topFrame = tk.Frame(self.root, width=self.width, height=self.topHeight)
        self.topFrame.pack(side=tk.TOP, padx=5)
        self.topFrame.pack_propagate(0)

        self.bottomHeight = int(self.height * 0.4)

        self.bottomFrame = tk.Frame(self.root, width=self.width, height=self.bottomHeight)
        self.bottomFrame.pack(side=tk.BOTTOM, padx=5)
        self.bottomFrame.pack_propagate(0)

    # ui layout
    def frameLayout(self):
        self.buttonLayout()
        self.argLayout()
        self.logLayout()
        self.resultLayout()

    # 按鈕 layout
    def buttonLayout(self):
        self.btnFrame = tk.LabelFrame(self.topFrame, text='功能', width=int(self.width * 0.25), height=self.topHeight)
        self.btnFrame.pack(side=tk.LEFT)
        self.btnFrame.pack_propagate(0)

        btn = tk.Button(self.btnFrame, text='爬', command=lambda: self.switchBtn(self.dataButonGroup))
        btn.place(x=5, y=5)

        btn = tk.Button(self.btnFrame, text='auto', command=lambda: self.switchBtn(self.auto_button_group))
        btn.place(x=self.w * 5, y=5)

        btn = tk.Button(self.btnFrame, text='cmoney', command=lambda: self.switchBtn(self.cmoneyButtonGroup))
        btn.place(x=self.w * 10, y=5)

        btn = tk.Button(self.btnFrame, text='個股', command=lambda: self.switchBtn(self.stockButtonGroup))
        btn.place(x=self.w * 18, y=5)

        btn = tk.Button(self.btnFrame, text='其他', command=lambda: self.switchBtn(self.otherButtonGroup))
        btn.place(x=5, y=self.h * 6)

        btn = tk.Button(self.btnFrame, text='xq', command=lambda: self.switchBtn(self.xqButtonGroup))
        btn.place(x=self.w * 5, y=self.h * 6)

        btn = tk.Button(self.btnFrame, text='execl', command=lambda: self.switchBtn(self.execlButtonGroup))
        btn.place(x=self.w * 10, y=self.h * 6)

        self.btnGroupFrame = tk.Frame(self.btnFrame, width=int(self.width * 0.25), bg='#eeeeee',
                                      height=int(self.topHeight * 0.7))
        self.btnGroupFrame.pack(side=tk.BOTTOM)
        self.btnGroupFrame.pack_propagate(0)

    # 參數 layout
    def argLayout(self):
        self.argFrame = tk.LabelFrame(
            self.topFrame,
            text='參數',
            bg='#eeeeee',
            width=self.width - int(self.width * 0.25),
            height=self.topHeight
        )

        self.argFrame.pack(side=tk.RIGHT)
        self.argFrame.pack_propagate(0)

    # log layout
    def logLayout(self):
        width = {
            1920: 77,
            3840: 100,
        }

        self.scrollbar = tk.Scrollbar(self.bottomFrame)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.listbox = tk.Listbox(
            self.bottomFrame,
            bg='#eeeeee',
            yscrollcommand=self.scrollbar.set,
            width=width[self.size.width]
        )

        self.listbox.pack(side=tk.RIGHT, fill=tk.BOTH)
        self.scrollbar.config(command=self.listbox.yview)

    # result layout
    def resultLayout(self):
        self.resultFrame = tk.Frame(self.bottomFrame, bg='#eeeeee', width=int(self.width * 0.25),
                                    height=self.bottomHeight)
        self.resultFrame.pack(side=tk.LEFT, pady=2)
        self.resultFrame.pack_propagate(0)

        btn = tk.Button(self.resultFrame, text='置頂', command=lambda: self.setWinTop())
        btn.place(x=5, y=5)

    # 抓取資料功能按鈕組群
    def dataButonGroup(self):
        btn = tk.Button(self.btnGroupFrame, text='trend', command=lambda: self.switchArg(cmoney.stock))
        btn.place(x=5, y=5)

        btn = tk.Button(self.btnGroupFrame, text='市場', command=lambda: self.switchArg(cmoney.market))
        btn.place(x=5, y=self.h * 6)

        btn = tk.Button(self.btnGroupFrame, text='個股產業股', command=lambda: self.switchArg(cmoney.StockIndustry))
        btn.place(x=5, y=self.h * 12)

        btn = tk.Button(self.btnGroupFrame, text='個股概念股', command=lambda: self.switchArg(cmoney.StockConcept))
        btn.place(x=5, y=self.h * 18)

        btn = tk.Button(self.btnGroupFrame, text='月營收', command=lambda: self.switchArg(twse.MonthRevenue))
        btn.place(x=5, y=self.h * 24)

        btn = tk.Button(self.btnGroupFrame, text='資產負債表', command=lambda: self.switchArg(twse.BalanceSheet))
        btn.place(x=5, y=self.h * 30)

        btn = tk.Button(self.btnGroupFrame, text='綜合損益表',
                        command=lambda: self.switchArg(twse.ConsolidatedIncomeStatement))
        btn.place(x=5, y=self.h * 36)

        btn = tk.Button(self.btnGroupFrame, text='現金流量表',
                        command=lambda: self.switchArg(twse.CashFlowStatement))
        btn.place(x=self.h * 15, y=5)

        btn = tk.Button(self.btnGroupFrame, text='權益變動表',
                        command=lambda: self.switchArg(twse.ChangesInEquity))
        btn.place(x=self.h * 15, y=self.h * 6)

        btn = tk.Button(self.btnGroupFrame, text='股利',
                        command=lambda: self.switchArg(twse.Dividend))
        btn.place(x=self.h * 15, y=self.h * 12)

        self.setLog('data')

    # 自動化功能按鈕組群
    def auto_button_group(self):
        # btn = tk.Button(self.btnGroupFrame, text='xq當日走勢與K截圖', command=lambda: self.switchArg(xq.stockImageDay))
        # btn.place(x=5, y=5)
        #
        # btn = tk.Button(self.btnGroupFrame, text='xq歷史走勢與K截圖', command=lambda: self.switchArg(xq.stockImageHistory))
        # btn.place(x=self.w * 13, y=5)
        #
        # btn = tk.Button(self.btnGroupFrame, text='xq大盤截圖', command=lambda: self.switchArg(xq.marketImage))
        # btn.place(x=5, y=self.h * 6)
        #
        # btn = tk.Button(self.btnGroupFrame, text='xq歷史大盤截圖', command=lambda: self.switchArg(xq.marketImageHistory))
        # btn.place(x=self.w * 8, y=self.h * 6)

        btn = tk.Button(self.btnGroupFrame, text='xq定位', command=lambda: self.switchArg(xq.move))
        btn.place(x=5, y=5)

        btn = tk.Button(self.btnGroupFrame, text='cmoney tick', command=lambda: self.switchArg(cmoney.Tick))
        btn.place(x=self.w * 8, y=5)

        self.setLog('auto')

    # cmoney功能按鈕組群
    def cmoneyButtonGroup(self):
        tk.Button(
            self.btnGroupFrame,
            text='日轉csv',
            command=lambda: self.switchArg(cmoney.dayToData)
        ).place(x=5, y=5)

        tk.Button(
            self.btnGroupFrame,
            text='年轉csv',
            command=lambda: self.switchArg(cmoney.yearToData)
        ).place(x=5, y=self.h * 6)

        tk.Button(
            self.btnGroupFrame,
            text='個股轉csv',
            command=lambda: self.switchArg(cmoney.stockToData)
        ).place(x=5, y=self.h * 12)

        tk.Button(
            self.btnGroupFrame,
            text='個股trend轉csv',
            command=lambda: self.switchArg(cmoney.StockTrendToCsv)
        ).place(x=5, y=self.h * 18)

        tk.Button(
            self.btnGroupFrame,
            text='市場trend轉csv',
            command=lambda: self.switchArg(cmoney.MarKetTrendToCsv)
        ).place(x=5, y=self.h * 24)

        tk.Button(
            self.btnGroupFrame,
            text='個股tick轉csv',
            command=lambda: self.switchArg(cmoney.StockTickToCsv)
        ).place(x=5, y=self.h * 30)

        tk.Button(
            self.btnGroupFrame,
            text='轉成一般年',
            command=lambda: self.switchArg(cmoney.YearToYear)
        ).place(x=5, y=self.h * 36)

        self.setLog('cmoney')

    # 選股功能按鈕組
    def stockButtonGroup(self):
        btn = tk.Button(self.btnGroupFrame, text='選股', command=lambda: self.switchArg(stock.select))
        btn.place(x=5, y=5)

        self.setLog('stock')

    # 其他功能按鈕組群
    def otherButtonGroup(self):
        btn = tk.Button(self.btnGroupFrame, text='個股資料轉json', command=lambda: self.switchArg(other.stockInfo))
        btn.place(x=5, y=5)

        self.setLog('other')

    # xq功能按鈕組群
    def xqButtonGroup(self):
        btn = tk.Button(self.btnGroupFrame, text='轉年', command=lambda: self.switchArg(xq.ToYear))
        btn.place(x=5, y=5)

        self.setLog('sq')

    # execl功能按鈕組群
    def execlButtonGroup(self):
        btn = tk.Button(self.btnGroupFrame, text='csv', command=lambda: self.switchArg(execl.ToCsv))
        btn.place(x=5, y=5)

        btn = tk.Button(self.btnGroupFrame, text='合併資產負債表', command=lambda: self.switchArg(twse.MergeBalanceSheet))
        btn.place(x=5, y=self.h * 6)

        btn = tk.Button(self.btnGroupFrame, text='合併綜合損益表',
                        command=lambda: self.switchArg(twse.MergeConsolidatedIncomeStatement))
        btn.place(x=5, y=self.h * 12)

        btn = tk.Button(self.btnGroupFrame, text='合併現金流量表', command=lambda: self.switchArg(twse.MergeCashFlowStatement))
        btn.place(x=5, y=self.h * 18)

        btn = tk.Button(self.btnGroupFrame, text='合併權益變動表', command=lambda: self.switchArg(twse.MergeChangesInEquity))
        btn.place(x=5, y=self.h * 24)

        btn = tk.Button(self.btnGroupFrame, text='合併月收盤價', command=lambda: self.switchArg(twse.MergeMonthClosePrice))
        btn.place(x=5, y=self.h * 30)

        btn = tk.Button(self.btnGroupFrame, text='合併財報', command=lambda: self.switchArg(twse.MergeFinancial))
        btn.place(x=5, y=self.h * 36)

        btn = tk.Button(self.btnGroupFrame, text='合併月營收', command=lambda: self.switchArg(twse.MergeMonthRevenue))
        btn.place(x=self.h * 20, y=5)

        btn = tk.Button(self.btnGroupFrame, text='合併股利', command=lambda: self.switchArg(twse.MergeDividend))
        btn.place(x=self.h * 20, y=self.h * 6)

        btn = tk.Button(self.btnGroupFrame, text='一鍵合併財報', command=lambda: self.switchArg(twse.MergeFinancials))
        btn.place(x=self.h * 20, y=self.h * 12)

        self.setLog('execl')

    # 切換 按鈕群組 layout內容
    def switchBtn(self, pack):
        self.listbox.delete(0, tk.END)

        for f in self.btnGroupFrame.winfo_children():
            f.destroy()

        pack()

        self.btnGroupFrame.pack()

    # 切換 參數 layout內容
    def switchArg(self, frame):
        self.listbox.delete(0, tk.END)

        for f in self.argFrame.winfo_children():
            f.destroy()

        frame(self.root, self.argFrame, self.w, self.h, self.config)

    # 視窗置頂
    def setWinTop(self):
        if self.isTop == False:
            self.isTop = True
        else:
            self.isTop = False

        self.root.wm_attributes('-topmost', self.isTop)

    def setLog(self, name):
        log.init(self.currentPath, name, self.listbox)


class image():
    def __init__(self, root, config=None):
        self.root = root
        self.size = pyautogui.size()
        self.width = self.size.width
        self.height = self.size.height
        self.w = self.width / 100
        self.h = self.height / 100
        self.config = config
        self.stock = {}
        self.img = {}

        root.geometry(f'{self.width}x{self.height}')

        self.mainLayout()
        self.functionLayout()
        self.modelFrameLayout()
        self.buttnoGroupLayout()

    def mainLayout(self):
        self.topHeight = int(self.height * 0.75)

        self.viewFrame = tk.Frame(self.root, width=self.width, height=self.topHeight, bg='grey')
        self.viewFrame.pack(side=tk.TOP)
        self.viewFrame.pack_propagate(0)

        self.bottomHeight = int(self.height * 0.25)

        self.functionFrame = tk.Frame(self.root, width=self.width, height=self.bottomHeight, bg='black')
        self.functionFrame.pack(side=tk.BOTTOM)
        self.functionFrame.pack_propagate(0)

    def functionLayout(self):
        self.modelFrame = tk.Frame(self.functionFrame, width=int((self.width / 3) / 2), height=self.bottomHeight)
        self.modelFrame.pack(side=tk.LEFT, padx=5)
        self.modelFrame.pack_propagate(0)

        self.stockFrame = tk.Frame(self.functionFrame, width=int((self.width / 3) / 2), height=self.bottomHeight)
        self.stockFrame.pack(side=tk.LEFT, padx=5)
        self.stockFrame.pack_propagate(0)

        self.inputFrame = tk.Frame(self.functionFrame, width=int(self.width / 3), height=self.bottomHeight)
        self.inputFrame.pack(side=tk.LEFT, padx=5)
        self.inputFrame.pack_propagate(0)

        self.groupListFrame = tk.Frame(self.functionFrame, width=int((self.width / 3) * 0.4), height=self.bottomHeight)
        self.groupListFrame.pack(side=tk.LEFT, padx=5)
        self.groupListFrame.pack_propagate(0)

        self.groupScrollbar = tk.Scrollbar(self.groupListFrame)
        self.groupScrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.groupListbox = tk.Listbox(
            self.groupListFrame,
            bg='#eeeeee',
            font=ui.BIG_FONT,
            selectbackground="orange",
            yscrollcommand=self.groupScrollbar.set
        )

        self.groupListbox.bind('<KeyRelease-Up>', self.groupListEvent)
        self.groupListbox.bind('<KeyRelease-Down>', self.groupListEvent)
        self.groupListbox.bind('<Button-1>', self.groupListEvent)
        self.groupListbox.pack(side=tk.RIGHT, fill=tk.BOTH)
        self.groupScrollbar.config(command=self.groupListbox.yview)

        self.stockListFrame = tk.Frame(self.functionFrame, width=int((self.width / 3) * 0.6), height=self.bottomHeight)
        self.stockListFrame.pack(side=tk.LEFT, padx=5)
        self.stockListFrame.pack_propagate(0)

        self.stockScrollbar = tk.Scrollbar(self.stockListFrame)
        self.stockScrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.stockListbox = tk.Listbox(
            self.stockListFrame,
            bg='#eeeeee',
            font=ui.BIG_FONT,
            selectbackground="orange",
            yscrollcommand=self.stockScrollbar.set,
            width=50
        )

        self.stockListbox.bind('<KeyRelease-Up>', self.stockListEvent)
        self.stockListbox.bind('<KeyRelease-Down>', self.stockListEvent)
        self.stockListbox.bind('<Button-1>', self.stockListEvent)
        self.stockListbox.pack(side=tk.RIGHT, fill=tk.BOTH)
        self.stockScrollbar.config(command=self.stockListbox.yview)

    def modelFrameLayout(self):
        self.kModel = tk.BooleanVar()
        self.trendModel = tk.BooleanVar()

        tk.Checkbutton(
            self.modelFrame,
            text='K',
            font=ui.FONT,
            var=self.kModel,
            command=lambda: self.setModel(ui.K, self.kModel.get())
        ).place(x=5, y=5)

        tk.Checkbutton(
            self.modelFrame,
            text='走勢',
            font=ui.FONT,
            var=self.trendModel,
            command=lambda: self.setModel(ui.TREND, self.trendModel.get())
        ).place(x=5, y=self.w * 2)

    def buttnoGroupLayout(self):
        self.path = tk.StringVar()
        self.dir = tk.StringVar()

        tk.Button(
            self.stockFrame,
            text='選擇目錄',
            font=ui.FONT,
            command=self.openGroupList
        ).place(x=5, y=5)

        tk.Label(self.stockFrame, textvariable=self.path, font=ui.FONT).place(x=5, y=self.w * 2)
        tk.Label(self.stockFrame, textvariable=self.dir, font=ui.FONT).place(x=5, y=self.w * 4)

    def openGroupList(self):
        self.groupListbox.delete(0, tk.END)

        path = ui.openDir()
        self.dir.set(os.path.split(path)[1])
        self.path.set(path)

        for f in os.listdir(path):
            f = os.path.splitext(f)

            if f[-1] == '.xlsx':
                self.groupListbox.insert(tk.END, f[0])

    def setModel(self, name, isDisplay):
        for w in self.viewFrame.winfo_children():
            w.destroy()

        if isDisplay:
            self.img[name] = ''
        elif name in self.img:
            del self.img[name]

        w = self.sizes()
        i = self.img.__len__() - 1

        for name in self.img:
            photo = ImageTk.PhotoImage(Image.new('RGB', w[i], color='grey'))
            self.img[name] = tk.Label(self.viewFrame, image=photo)
            self.img[name].image = photo

            if self.img.__len__() == 1:
                self.img[name].pack(padx=5, pady=5)
            else:
                self.img[name].pack(side=tk.LEFT, padx=5, pady=5)

        self.setCurCode()

    def groupListEvent(self, event):
        self.stockListbox.delete(0, tk.END)
        self.stock.clear()

        path = os.path.join(self.path.get(), str(self.groupListbox.get(tk.ACTIVE))) + '.xlsx'
        ws = openpyxl.load_workbook(path)
        sheet = ws.active

        for rows in sheet.iter_rows(2, 0, 0, sheet.max_column):
            self.stockListbox.insert(tk.END, f'{rows[0].value} - {rows[1].value}')

        imageDir = os.path.join(self.path.get(), 'image', self.groupListbox.get(tk.ACTIVE))

        for d in os.listdir(imageDir):
            for f in glob.glob(os.path.join(imageDir, d, '*.png')):
                code = os.path.basename(f).split('.')[0]

                if code in self.stock:
                    self.stock[code][d] = f
                else:
                    self.stock[code] = {
                        d: f,
                    }

    def stockListEvent(self, event):
        self.setCurCode()

    def setCurCode(self):
        name = self.stockListbox.get(tk.ACTIVE)
        self.setCode(name.split('-')[0].strip())

    def setCode(self, code):
        if code in self.stock:
            for name in self.img:
                img = ImageTk.PhotoImage(
                    Image.open(self.stock[code][name]).resize(
                        (self.img[name].image.width(), self.img[name].image.height())
                    )
                )

                self.img[name].configure(image=img)
                self.img[name].image = img

    def sizes(self):
        return [(int(self.width * 0.6), self.topHeight), (int(self.width / 2), self.topHeight)]


class Watch():
    default_code = 2330

    def __init__(self, root, config=None, path=None):
        self.root = root
        self.size = pyautogui.size()
        self.width = self.size.width
        self.height = self.size.height
        self.dir = ''
        self.type = None
        self.marker_dates = []

        # self.root.attributes('-fullscreen', True)
        self.config = config
        self.plot_config = dict(
            volume=True,
            max_min=True,
            max_min_text=True,
            avg=True,
            panel_ratios=(4, 1),
            ma=[5, 10, 20, 60, 120],
        )

        self.root.state('zoomed')
        self._mainLayout()

        self.watch = watch.Watch(
            self.watch_frame,
            width=self.width,
            height=self.height,
            config=config,
        )

        self._button_layout()
        self._list_layout()

        # self._plot_k_trend()
        # self._plot_trend()
        self._plot_k()
        self.watch.pack()

        filename = os.path.join(path, datetime.now().strftime(f"%Y-%m-%d-watch.log"))
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s [%(levelname)s] %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S',
            filename=filename
        )

    def _mainLayout(self):
        self.watch_frame = tk.Frame(self.root, width=int(self.width * 0.9), height=self.height)
        self.watch_frame.pack(side=tk.LEFT)
        self.watch_frame.pack_propagate(0)

        self.right_width = int(self.width * 0.1)
        self.right_frame = tk.Frame(self.root, width=self.right_width, height=self.height, bg='#C0C0C0')
        self.right_frame.pack(side=tk.RIGHT)
        self.right_frame.pack_propagate(0)

        self.top_frame = tk.Frame(self.right_frame, width=self.right_width, height=int(self.height * 0.5))
        self.top_frame.pack(side=tk.TOP)
        self.top_frame.pack_propagate(0)

        self.bottom_frame = tk.Frame(self.right_frame, width=self.right_width, height=int(self.height * 0.5),
                                     bg='#E0E0E0')
        self.bottom_frame.pack(side=tk.BOTTOM)
        self.bottom_frame.pack_propagate(0)

    def _button_layout(self):
        self.code = tk.StringVar()
        self.date = tk.StringVar()
        self.start_date_range = tk.StringVar()
        self.end_date_range = tk.StringVar()
        self.type_name = tk.StringVar()

        self.code.set(self.default_code)

        tk.Label(self.bottom_frame, text='個股:', font=ui.FONT).place(x=10, y=10)
        tk.Entry(self.bottom_frame, width=8, textvariable=self.code, font=ui.FONT).place(x=130, y=10)
        tk.Label(self.bottom_frame, text='日期:', font=ui.FONT).place(x=10, y=100)
        tk.Entry(self.bottom_frame, width=10, textvariable=self.date, font=ui.BTN_FONT).place(x=130, y=100)
        tk.Label(self.bottom_frame, text='分類:', font=ui.FONT).place(x=10, y=200)
        tk.Entry(self.bottom_frame, width=10, textvariable=self.type_name, font=ui.BTN_FONT).place(x=130, y=200)
        tk.Label(self.bottom_frame, text='sx:', font=ui.FONT).place(x=10, y=300)
        tk.Entry(self.bottom_frame, width=10, textvariable=self.start_date_range, font=ui.BTN_FONT).place(x=130, y=300)
        tk.Label(self.bottom_frame, text='ex:', font=ui.FONT).place(x=10, y=400)
        tk.Entry(self.bottom_frame, width=10, textvariable=self.end_date_range, font=ui.BTN_FONT).place(x=130, y=400)

        tk.Button(
            self.bottom_frame,
            text='切',
            font=ui.SMALL_FONT,
            command=self._update_plot,
        ).place(x=10, y=500)

        tk.Button(
            self.bottom_frame,
            text='K',
            font=ui.SMALL_FONT,
            command=self._plot_k,
        ).place(x=100, y=500)
        tk.Button(
            self.bottom_frame,
            text='勢',
            font=ui.SMALL_FONT,
            command=self._plot_trend,
        ).place(x=180, y=500)
        tk.Button(
            self.bottom_frame,
            text='k勢',
            font=ui.SMALL_FONT,
            command=self._plot_k_trend,
        ).place(x=270, y=500)
        tk.Button(
            self.bottom_frame,
            text='載',
            font=ui.SMALL_FONT,
            command=self._open_dir_stock,
        ).place(x=10, y=580)
        tk.Button(
            self.bottom_frame,
            text='類',
            font=ui.SMALL_FONT,
            command=self._save_code,
        ).place(x=100, y=580)

    def _list_layout(self):
        date_list = tk.Frame(self.top_frame, width=self.right_width, height=int(self.height * 0.15))
        date_list.pack(side=tk.TOP)
        date_list.pack_propagate(0)

        stock_list = tk.Frame(self.top_frame, width=self.right_width, height=int(self.height * 0.35))
        stock_list.pack(side=tk.TOP)
        stock_list.pack_propagate(0)

        date_Scrollbar = tk.Scrollbar(date_list)
        date_Scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self._date_listbox = tk.Listbox(
            date_list,
            bg='#eeeeee',
            font=ui.FONT,
            selectbackground="orange",
            yscrollcommand=date_Scrollbar.set,
        )

        stock_Scrollbar = tk.Scrollbar(stock_list)
        stock_Scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self._stock_listbox = tk.Listbox(
            stock_list,
            bg='#eeeeee',
            font=ui.FONT,
            selectbackground="orange",
            yscrollcommand=stock_Scrollbar.set,
        )

        self._date_listbox.bind('<KeyRelease-Up>', self._date_event)
        self._date_listbox.bind('<KeyRelease-Down>', self._date_event)
        self._date_listbox.bind('<Button-1>', self._date_event)

        self._stock_listbox.bind('<KeyRelease-Up>', self._stock_event)
        self._stock_listbox.bind('<KeyRelease-Down>', self._stock_event)
        self._stock_listbox.bind('<Button-1>', self._stock_event)

        date_Scrollbar.config(command=self._date_listbox.yview)
        stock_Scrollbar.config(command=self._stock_listbox.yview)
        self._date_listbox.pack(side=tk.RIGHT, fill=tk.BOTH)
        self._stock_listbox.pack(side=tk.RIGHT, fill=tk.BOTH)

    def _open_dir_stock(self):
        self._date_list = {}
        self.dir = ui.openDir()
        self._date_listbox.delete(0, tk.END)
        self._stock_listbox.delete(0, tk.END)

        for path in sorted(glob.glob(os.path.join(self.dir, '*.csv')), reverse=True):
            name = os.path.basename(path).split('.')[0]
            self._date_listbox.insert(tk.END, name)
            self._date_list[name] = pd.read_csv(path)

    def _date_event(self, event):
        self._stock_listbox.delete(0, tk.END)

        date = self._date_listbox.get(tk.ACTIVE)

        self.date.set(date)

        for index, row in self._date_list[date].iterrows():
            self._stock_listbox.insert(tk.END, f"{row['code']}-{row['name']}")

    def _stock_event(self, event):
        value = self._stock_listbox.get(tk.ACTIVE)
        code = value.split('-')[0]
        data = self._date_list[self.date.get()]
        data = data[data['code'] == int(code)]

        self.code.set(code)

        if 'start_date' in data.columns:
            self.start_date_range.set(data['start_date'].iloc[0])
        else:
            self.start_date_range.set('')

        if 'end_date' in data.columns:
            self.end_date_range.set(data['end_date'].iloc[0])
        else:
            self.end_date_range.set('')

        self.marker_dates = []
        for c in ['right_date', 'left_date', 'left_bottom_date']:
            if c in data:
                self.marker_dates.append(data[c].iloc[0])

        if self.type == 'k':
            self._plot_k()
        elif self.type == 'trend':
            self._plot_trend()
        elif self.type == 'k_trend':
            self._plot_k_trend()

    def _plot_k(self):
        self._plot('k')

    def _plot_trend(self):
        self._plot('trend')

    def _plot_k_trend(self):
        self._plot('k_trend')

    def _plot(self, type):
        self.type = type
        self._update_config()

        date = self.date.get()
        if date == '':
            date = None

        self.watch.plot(int(self.code.get()), date=date, type=type, **self.plot_config)

    def _update_plot(self):
        self._update_config()

        if self.watch.update_plot(int(self.code.get()), date=self.date.get(), type=self.type,
                                  **self.plot_config) == False:
            messagebox.showinfo('結果', '無此個股')

    def _update_config(self):
        self.plot_config['range'] = [self.start_date_range.get(), self.end_date_range.get()]
        self.plot_config['marker_date'] = self.marker_dates

    def _save_code(self):
        name = self.type_name.get()
        if (name is None) or (name == ''):
            return

        dir = os.path.join(self.dir, name)

        if os.path.exists(dir) == False:
            os.mkdir(dir)

        data = pd.DataFrame(columns=['date', 'code'])
        date = self.date.get()
        file = os.path.join(dir, date) + '.csv'

        if os.path.exists(file):
            data = pd.read_csv(file)

        data = data.append([{'date': date, 'code': self.code.get()}])

        data.to_csv(file, index_label=['date', 'code'], index=False)


class XQ():
    def __init__(self, root, config=None, path=None):
        self.root = root
        self.size = pyautogui.size()
        self.width = self.size.width
        self.height = self.size.height
        self.dir = ''
        self.config = config
        self.stock = data.Stock(other.stock_csv_path(config))

        self.root.state('zoomed')
        self._mainLayout()
        self._button_layout()
        self._list_layout()

        filename = os.path.join(path, datetime.now().strftime(f"%Y-%m-%d-watch.log"))
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s [%(levelname)s] %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S',
            filename=filename
        )

    def _mainLayout(self):
        self.right_width = int(self.width * 0.1)
        self.right_frame = tk.Frame(self.root, width=self.right_width, height=self.height, bg='#C0C0C0')
        self.right_frame.pack(side=tk.LEFT)
        self.right_frame.pack_propagate(0)

        self.top_frame = tk.Frame(self.right_frame, width=self.right_width, height=int(self.height * 0.5))
        self.top_frame.pack(side=tk.TOP)
        self.top_frame.pack_propagate(0)

        self.bottom_frame = tk.Frame(self.right_frame, width=self.right_width, height=int(self.height * 0.5),
                                     bg='#E0E0E0')
        self.bottom_frame.pack(side=tk.BOTTOM)
        self.bottom_frame.pack_propagate(0)

    def _button_layout(self):
        self.code = tk.StringVar()
        self.date = tk.StringVar()
        self.start_date = tk.StringVar()
        self.now_date = tk.StringVar()

        tk.Label(self.bottom_frame, text='個股:', font=ui.FONT).place(x=10, y=10)
        tk.Entry(self.bottom_frame, width=8, textvariable=self.code, font=ui.FONT).place(x=130, y=10)
        tk.Label(self.bottom_frame, text='日期:', font=ui.FONT).place(x=10, y=100)
        tk.Entry(self.bottom_frame, width=10, textvariable=self.date, font=ui.BTN_FONT).place(x=130, y=100)
        tk.Label(self.bottom_frame, text='現年月:', font=ui.FONT).place(x=10, y=200)
        tk.Entry(self.bottom_frame, width=8, textvariable=self.start_date, font=ui.BTN_FONT).place(x=180, y=200)
        tk.Label(self.bottom_frame, text='起年月:', font=ui.FONT).place(x=10, y=300)
        tk.Entry(self.bottom_frame, width=8, textvariable=self.now_date, font=ui.BTN_FONT).place(x=180, y=300)

        tk.Button(
            self.bottom_frame,
            text='載',
            font=ui.SMALL_FONT,
            command=self._open_dir_stock,
        ).place(x=10, y=400)

        tk.Button(
            self.bottom_frame,
            text='切',
            font=ui.SMALL_FONT,
            command=lambda: self._update(self.code.get()),
        ).place(x=100, y=400)

    def _list_layout(self):
        date_list = tk.Frame(self.top_frame, width=self.right_width, height=int(self.height * 0.15))
        date_list.pack(side=tk.TOP)
        date_list.pack_propagate(0)

        stock_list = tk.Frame(self.top_frame, width=self.right_width, height=int(self.height * 0.35))
        stock_list.pack(side=tk.TOP)
        stock_list.pack_propagate(0)

        date_Scrollbar = tk.Scrollbar(date_list)
        date_Scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self._date_listbox = tk.Listbox(
            date_list,
            bg='#eeeeee',
            font=ui.FONT,
            selectbackground="orange",
            yscrollcommand=date_Scrollbar.set,
        )

        stock_Scrollbar = tk.Scrollbar(stock_list)
        stock_Scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self._stock_listbox = tk.Listbox(
            stock_list,
            bg='#eeeeee',
            font=ui.FONT,
            selectbackground="orange",
            yscrollcommand=stock_Scrollbar.set,
        )

        self._date_listbox.bind('<KeyRelease-Up>', self._date_event)
        self._date_listbox.bind('<KeyRelease-Down>', self._date_event)
        self._date_listbox.bind('<Button-1>', self._date_event)

        self._stock_listbox.bind('<KeyRelease-Up>', self._stock_event)
        self._stock_listbox.bind('<KeyRelease-Down>', self._stock_event)
        self._stock_listbox.bind('<ButtonRelease-1>', self._stock_btn_event)

        date_Scrollbar.config(command=self._date_listbox.yview)
        stock_Scrollbar.config(command=self._stock_listbox.yview)
        self._date_listbox.pack(side=tk.RIGHT, fill=tk.BOTH)
        self._stock_listbox.pack(side=tk.RIGHT, fill=tk.BOTH)

    def _open_dir_stock(self):
        self._date_list = {}
        self.dir = ui.openDir()
        self._date_listbox.delete(0, tk.END)
        self._stock_listbox.delete(0, tk.END)

        for path in sorted(glob.glob(os.path.join(self.dir, '*.csv')), reverse=True):
            name = os.path.basename(path).split('.')[0]
            self._date_listbox.insert(tk.END, name)
            self._date_list[name] = pd.read_csv(path)

    def _date_event(self, event):
        self._stock_listbox.delete(0, tk.END)

        date = self._date_listbox.get(tk.ACTIVE)

        self.date.set(date)

        for index, row in self._date_list[date].iterrows():
            self._stock_listbox.insert(tk.END, f"{row['code']}-{row['name']}")

    def _stock_btn_event(self, event):
        time.sleep(1)
        self._stock_event(event)

    def _stock_event(self, event):
        value = self._stock_listbox.get(event.widget.curselection()[0])
        code = value.split('-')[0]
        self.code.set(code)
        self._update(code)

    def _update(self, code):
        pyautogui.click(1800, 90)
        pyautogui.write(str(code))
        pyautogui.press('enter')

        nYear = None
        nMonth = None

        if self.now_date.get() != '':
            nYear = int(self.now_date.get()[:4])
            nMonth = int(self.now_date.get()[5:7])

        xqa.k().dates(self.stock.afterDates(self.date.get())[40], self.date.get(),
                      year=int(self.start_date.get()[:4]),
                      month=int(self.start_date.get()[5:7]),
                      nYear=nYear,
                      nMonth=nMonth
                      )

        pyautogui.click(3500, 1700)
