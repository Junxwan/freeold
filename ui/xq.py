# -*- coding: utf-8 -*-

import os
import tkinter as tk
import openpyxl
from . import ui, other
from auto import xq as xq
from stock import data
from xlsx import xq as execl


class stockImage(ui.process):
    def __init__(self, root, master, w, h, config=None):
        ui.process.__init__(self, master, w, h)

        self.code = tk.StringVar()
        self.config = config

        tk.Label(master, text='代碼:', font=ui.FONT).place(x=10, y=10)
        tk.Entry(master, textvariable=self.code, font=ui.FONT).place(x=self.ex, y=10)
        tk.Button(
            master,
            text='選擇檔案',
            font=ui.BTN_FONT,
            command=lambda: self.code.set(ui.openFile().name)
        ).place(x=self.w * 50, y=10)

        self.addRunBtn(master)

    def outputPath(self):
        return os.path.join(
            os.path.dirname(self.code.get()), 'image', os.path.basename(self.code.get()).split('.')[0]
        )

    def getCodes(self):
        codes = []
        xlsx = openpyxl.load_workbook(self.code.get())
        sheet = xlsx.active

        for rows in sheet.iter_rows(2, 0, 0, sheet.max_column):
            if rows[0].value == None:
                continue

            codes.append(rows[0].value)

        return codes


# auto 自動擷取當日走勢與技術分析圖參數
class stockImageDay(stockImage):
    def __init__(self, root, master, w, h, config=None):
        stockImage.__init__(self, root, master, w, h, config)

    def run(self):
        xq.stockNow().start(self.getCodes(), self.outputPath())


# auto 自動擷取歷史走勢與技術分析圖參數
class stockImageHistory(stockImage):
    def __init__(self, root, master, w, h, config=None):
        stockImage.__init__(self, root, master, w, h, config)

        self.dir = tk.StringVar()
        self.date = tk.StringVar()

        if config != None:
            self.dir.set(config['json'])

        tk.Label(master, text='檔案:', font=ui.FONT).place(x=10, y=self.ey)
        tk.Entry(master, textvariable=self.dir, font=ui.FONT).place(x=self.ex, y=self.ey)
        tk.Button(
            master,
            text='選擇目錄',
            font=ui.BTN_FONT,
            command=lambda:
            self.dir.set(ui.openDir())
        ).place(x=self.w * 50, y=self.h * 8)

    def run(self):
        xq.stockHistory(os.path.basename(self.code.get()).split('.')[0], self.dir.get()).start(
            self.getCodes(),
            self.outputPath()
        )


# auto 大盤圖參數
class marketImage(ui.process):
    def __init__(self, root, master, w, h, config=None):
        ui.process.__init__(self, master, w, h)

        self.output = tk.StringVar()

        if config != None:
            self.output.set(config['output'])

        tk.Label(master, text='輸出:', font=ui.FONT).place(x=10, y=10)
        tk.Entry(master, textvariable=self.output, font=ui.FONT).place(x=self.ex, y=10)
        tk.Button(
            master,
            text='選擇目錄',
            font=ui.BTN_FONT,
            command=lambda: self.output.set(ui.openDir())
        ).place(x=self.w * 50, y=10)

        self.addRunBtn(master)

    def run(self):
        xq.marketNow().start(self.output.get())


# auto 歷史大盤圖參數
class marketImageHistory(ui.process):
    def __init__(self, root, master, w, h, config=None):
        ui.process.__init__(self, master, w, h)

        self.dir = tk.StringVar()
        self.date = tk.StringVar()
        self.output = tk.StringVar()

        if config != None:
            self.dir.set(config['json'])
            self.output.set(config['output'])

        tk.Label(master, text='日期:', font=ui.FONT).place(x=10, y=10)
        tk.Entry(master, textvariable=self.date, font=ui.FONT).place(x=self.ex, y=10)

        tk.Label(master, text='檔案:', font=ui.FONT).place(x=10, y=self.ey)
        tk.Entry(master, textvariable=self.dir, font=ui.FONT).place(x=self.ex, y=self.ey)
        tk.Button(
            master,
            text='選擇目錄',
            font=ui.BTN_FONT,
            command=lambda:
            self.dir.set(ui.openDir())
        ).place(x=self.w * 50, y=self.h * 8)

        tk.Label(master, text='輸出:', font=ui.FONT).place(x=10, y=self.ey * 2)
        tk.Entry(master, textvariable=self.output, font=ui.FONT).place(x=self.ex, y=self.ey * 2)
        tk.Button(
            master,
            text='選擇目錄',
            font=ui.BTN_FONT,
            command=lambda: self.output.set(ui.openDir())
        ).place(x=self.w * 50, y=self.h * 18)

        self.addRunBtn(master)

    def run(self):
        xq.marketHistory(self.date.get(), self.dir.get()).start(self.output.get())


class move(ui.process):
    def __init__(self, root, master, w, h, config=None):
        ui.process.__init__(self, master, w, h)

        self.start_now_date = tk.StringVar()
        self.end_date = tk.StringVar()
        self.stock = data.Stock(other.stock_csv_path(config))

        tk.Label(master, text='日期:', font=ui.FONT).place(x=10, y=10)
        tk.Entry(master, textvariable=self.end_date, font=ui.FONT).place(x=self.ex * 1.5, y=10)

        tk.Label(master, text='現在開始年月:', font=ui.FONT).place(x=10, y=self.ey)
        tk.Entry(master, textvariable=self.start_now_date, font=ui.FONT).place(x=self.ex * 1.5, y=self.ey)

        tk.Button(
            master,
            text='K',
            font=ui.BTN_FONT,
            command=lambda: self.k()
        ).place(x=10, y=self.h * 18)

        self.keyEvent(root)

    def k(self, event=None):
        xq.k().dates(self.stock.afterDates(self.end_date.get())[40], self.end_date.get(),
                     year=int(self.start_now_date.get()[:4]),
                     month=int(self.start_now_date.get()[5:7]))

    def keyEvent(self, master):
        master.focus_set()
        master.bind("<F1>", self.k)

class toData(ui.process):
    def __init__(self, root, master, w, h, config=None):
        ui.process.__init__(self, master, w, h)

        self.input = tk.StringVar()
        self.output = tk.StringVar()

        if config != None:
            self.output.set(os.path.join(config['data'], 'csv'))

        self.default_output()

        tk.Label(master, text='xlsx:', font=ui.FONT).place(x=10, y=10)
        tk.Entry(master, textvariable=self.input, font=ui.FONT).place(x=self.ex, y=10)
        tk.Button(
            master,
            text=self.openInputText(),
            font=ui.BTN_FONT,
            command=self.openInput
        ).place(x=self.w * 50, y=10)

        tk.Label(master, text='輸出:', font=ui.FONT).place(x=10, y=self.ey)
        tk.Entry(master, textvariable=self.output, font=ui.FONT).place(x=self.ex, y=self.ey)
        tk.Button(
            master,
            text='選擇目錄',
            font=ui.BTN_FONT,
            command=self.openOutPut
        ).place(x=self.w * 50, y=self.h * 8)

        self.addRunBtn(master)

    def default_output(self):
        pass

    def openInputText(self):
        return ""

    def openInput(self):
        pass

    def openOutPut(self):
        pass


class ToYear(toData):
    def openInputText(self):
        return '選擇檔案'

    def openInput(self):
        self.input.set(ui.openDir())

    def openOutPut(self):
        self.output.set(ui.openDir())

    def run(self):
        execl.ToYear(self.input.get(), self.output.get())
        self.showSuccess()